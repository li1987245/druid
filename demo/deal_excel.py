#!/usr/bin/python3
# -*- coding: utf-8 -*-
import os
import pickle

import requests
import xlrd
import xlwt
from openpyxl import load_workbook, Workbook
from requests.auth import HTTPBasicAuth
import MySQLdb as mdb


def read_excel():
    invoices = []
    invoice_comps = []
    readbook = xlrd.open_workbook(r'D:\document\工作\数据洞察\开票历史数据\18年开票明细原始.xlsx')
    # sheet = readbook.sheet_by_index(0)  # 索引的方式，从0开始
    sheet = readbook.sheet_by_name('18年开票')  # 名字的方式
    nrows = sheet.nrows  # 行
    ncols = sheet.ncols  # 列
    for i in range(nrows)[1:]:
        invoice_comps.append(sheet.cell(i, 0).value)
        invoices.append(sheet.row_values(i))
    # sheet.col_values(i)
    # sheet = readbook.sheet_by_name('18年期初累计未开票')  # 名字的方式
    # nrows = sheet.nrows  # 行
    # for i in range(nrows)[1:]:
    #     invoice_comps.append(sheet.cell(i, 0).value)
    #     invoices.append(sheet.row_values(i))
    return invoice_comps, invoices


def write_excel(invoices):
    workbook = xlwt.Workbook(encoding='ascii')
    worksheet = workbook.add_sheet('18年开票-含cid')
    for invoice in invoices:
        worksheet.write(0, 0, 'Row 0, Column 0 Value')
    workbook.save('Excel_Workbook.xls')


def rewrite_excel(cids, dic):
    workbook_ = load_workbook('18年开票明细.xlsx')
    # workbook_ = load_workbook('18年开票明细.xlsx')
    sheetnames = workbook_.get_sheet_names()  # 获得表单名字
    sheet = workbook_.get_sheet_by_name("18年开票")
    # sheet = workbook_.get_sheet_by_name("18年期初累计未开票")
    rows = sheet.rows
    columns = sheet.columns
    # sheet['A1'] = '47'
    sheet.cell(row=1, column=7, value="cid")
    for i, cid in enumerate(cids):
        if cid == '':
            iname = sheet.cell(row=i + 2, column=1).value
            cid = dic.get(iname)
        sheet.cell(row=i + 2, column=7, value=cid)
    workbook_.save(u"18年开票明细.xlsx")
    # wb = Workbook()
    # ws = wb.active
    # ws['A1'] = 4
    # wb.save("新歌检索失败.xlsx")


def req_crm(invoice_comps):
    """
    根据crm接口获取开票主体对应cid
    :param invoice_comps:
    :return:
    """
    cids = []
    if os.path.exists('tmp.pk'):
        with open('tmp.pk', 'rb') as f:
            cids = pickle.load(f)
        return cids
    url = 'http://crm.100credit.cn/api/customer/info'
    for k, v in enumerate(invoice_comps):
        dic = requests.get(url, params={"paymentAccount": v}).json()
        code = dic.get("code")
        compId = ''
        if code == 0:
            try:
                lst = dic.get("data").get("list")
                if len(lst) == 1:
                    compId = lst[0].get("compId")
            except Exception as e:
                print('Error:', e)
            finally:
                print(compId)
        cids.append(compId)
    # r = requests.post(url, data={"paymentAccount":invoice_comps[0]}, auth=HTTPBasicAuth('admin', 'admin'))
    # r = requests.post(url, data={"paymentAccount": invoice_comps[0]})
    # print(r.json())
    with open('tmp.pk', 'wb') as f:
        pickle.dump(cids, f)
    return cids


def op_mysql():
    dic = {}
    conn = mdb.connect(host='192.168.162.235', port=3306, user='bill', password='mysql', database='bill',
                       charset='utf8')
    # cursor = conn.cursor()
    # cursor.execute('insert into user (id, name) values (%s, %s)', ['1', 'Michael'])
    # cursor.rowcount
    # conn.commit()
    # cursor.close()
    cursor = conn.cursor()
    # cursor.execute('select * from user where id = %s', ('1',))
    cursor.execute('select * from T_ACCOUNT_BALANCE_18')
    values = cursor.fetchall()
    for value in values:
        iname = value[1]
        cid = value[7]
        # print('iname:%s cid:%s' % (iname, cid))
        dic[iname] = cid
    cursor.close()
    cursor = conn.cursor()
    # cursor.execute('select * from user where id = %s', ('1',))
    cursor.execute('select * from T_CONFIRM_AMOUNT_18')
    values = cursor.fetchall()
    for value in values:
        iname = value[3]
        cid = value[10]
        if cid is None or cid == '':
            print('iname:%s cid:%s' % (iname, cid))
            continue
        dic[iname] = cid
    cursor.close()
    conn.close()
    return dic


def insert_invoice_18():
    import re
    conn = mdb.connect(host='192.168.162.235', port=3306, user='bill', password='mysql', database='bill',
                       charset='utf8')
    readbook = xlrd.open_workbook(r'18年开票明细.xlsx')
    sheet = readbook.sheet_by_name('18年开票')  # 名字的方式
    nrows = sheet.nrows  # 行
    values = []
    for i in range(nrows)[1:]:
        iname = sheet.cell(i, 0).value
        sale = sheet.cell(i, 1).value
        type = sheet.cell(i, 2).value
        business_line_name = sheet.cell(i, 2).value + sheet.cell(i, 3).value
        business_line = foo(business_line_name)
        invoice_amount = str(sheet.cell(i, 5).value)
        cid = ''
        try:
            cid = str(int(sheet.cell(i, 6).value))
        except Exception as e:
            print(e)
        finally:
            pass
        values.append((iname, sale, type, business_line, business_line_name, invoice_amount, cid))
    cursor = conn.cursor()
    cursor.execute('truncate table T_ACCOUNT_INVOICE_18')
    cursor.executemany(
        'insert into T_ACCOUNT_INVOICE_18 (iname, sale,type,business_line,business_line_name,invoice_amount,cid) values (%s, %s, %s, %s, %s, %s, %s)',
        values)
    conn.commit()
    cursor.close()
    sheet = readbook.sheet_by_name('18年期初累计未开票')  # 名字的方式
    nrows = sheet.nrows  # 行
    values = []
    for i in range(nrows)[1:]:
        iname = sheet.cell(i, 0).value
        sale = sheet.cell(i, 1).value
        type = sheet.cell(i, 2).value
        business_line_name = sheet.cell(i, 2).value + sheet.cell(i, 3).value
        business_line = foo(business_line_name)
        invoice_amount = str(sheet.cell(i, 5).value)
        cid = ''
        try:
            cid = str(int(sheet.cell(i, 6).value))
        except Exception as e:
            print(e)
        finally:
            pass
        values.append((iname, sale, type,business_line,business_line_name, invoice_amount, cid))
    cursor = conn.cursor()
    cursor.execute('truncate table T_ACCOUNT_UNINVOICE_18')
    cursor.executemany(
        'insert into T_ACCOUNT_UNINVOICE_18 (iname, sale,type,business_line,business_line_name,invoice_amount,cid) values (%s, %s, %s, %s, %s, %s, %s)',
        values)
    conn.commit()
    cursor.close()

def foo(x):
    return {
        '资产管理催收':'A0403',
        '互联网榕树': 'A0201',
        '保险': 'A0301',
        '营销-华北营销-华北': 'A0201',
        '风控': 'A0101',
        '普惠金融': 'A0110',
        '互联网非榕树': 'A0201',
        '资产管理': 'A0403',
        '营销-华北': 'A0201',
    }.get(x)

dic = op_mysql()
invoice_comps, invoices = read_excel()
cids = req_crm(invoice_comps)
rewrite_excel(cids, dic)
# insert_invoice_18()
